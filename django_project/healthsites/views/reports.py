# -*- coding: utf-8 -*-
from __future__ import unicode_literals, print_function, absolute_import, division

import os
from django.contrib.auth.decorators import login_required
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponse

from openpyxl import load_workbook

from ..models.daily_report import DailyReport
from ..models.assessment import HealthsiteAssessment


@login_required
def reports(request):
    """View for request."""
    daily_reports = DailyReport.objects.exclude(assessment_number=0).order_by('-date_time')
    return render_to_response(
        'event_mapper/reports/reports_page.html',
        {
            'daily_reports': daily_reports
        },
        context_instance=RequestContext(request)

    )


def download_report(request, report_id):
    """The view to download users data as CSV.

    :param request: A django request object.
    :type request: request

    :return: A PDF File
    :type: HttpResponse
    """
    report = DailyReport.objects.get(id=report_id)
    fsock = open(report.file_path, 'r')
    response = HttpResponse(fsock, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="%s"' % (
        os.path.basename(report.file_path)
    )
    return response


def download_assessment_report(request, assessment_id):
    """The view to download users data as CSV.

    :param request: A django request object.
    :type request: request

    :return: A CSV File
    :type: HttpResponse
    """
    assessment = HealthsiteAssessment.objects.get(id=assessment_id)
    context = assessment.get_context_data()

    workbook = load_workbook('healthsites/templates/report.xlsx')
    sheet = workbook.get_sheet_by_name('0_SurveyList')
    for row in range(4, 65, 1):
        key = sheet.cell('J%s' % row).value
        if key in context:
            sheet.cell('J%s' % row).value = context[key]

    download_file_name = 'report_%s.xlsx' % assessment_id
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s' % download_file_name

    workbook.save(response)

    return response


def download_assessment_csv(request, assessment_id):
    import csv
    """The view to download users data as CSV.

    :param request: A django request object.
    :type request: request

    :return: A CSV File
    :type: HttpResponse
    """
    output = []
    try:
        event = HealthsiteAssessment.objects.get(id=assessment_id)
        dict = event.get_dict(True)
        for key in dict.keys():
            if key != 'assessment' and key != 'healthsite':
                output.append({'key': key, 'value': dict[key]})
        for key in dict['healthsite'].keys():
            if key == 'geometry':
                output.append({'key': 'latitude', 'value': dict['healthsite'][key][0]})
                output.append({'key': 'longitude', 'value': dict['healthsite'][key][1]})
            else:
                output.append({'key': key, 'value': dict['healthsite'][key]})
        for key in dict['assessment'].keys():
            output.append({'key': key, 'value': dict['assessment'][key]['description']})
    except HealthsiteAssessment.DoesNotExist:
        pass
    # create the csv writer object# Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='application/csv')
    response['Content-Disposition'] = 'attachment; filename="assessment_' + assessment_id + '.csv"'
    # convert to csv
    writer = csv.writer(response)
    writer.writerow([s['key'].encode('utf-8') if type(s['key']) is unicode else s['key'] for s in output])
    writer.writerow([s['value'].encode('utf-8') if type(s['value']) is unicode else s['value'] for s in output])
    return response
